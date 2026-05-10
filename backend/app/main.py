"""Overseas Relic Knowledge Service — FastAPI entrypoint.

Neo4j integration flow:
- ``NEO4J_URI`` set → ``/relics`` tries graph (see ``app.services.relics``); on failure or empty → JSON
- No URI → JSON only (local dev without Docker Neo4j)
"""

import csv
import html
import io
import json
import os
import re
from collections import Counter
from pathlib import Path
from typing import Any, Literal

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from app.services import relics as relics_service

# Repo root: backend/app/main.py -> parents[2] == Relic-system
REPO_ROOT = Path(__file__).resolve().parents[2]
SAMPLE_RELICS_PATH = REPO_ROOT / "database" / "graph" / "sample_relics.json"

EXPORT_MAX_ROWS = 10000

EXPORT_CSV_COLUMNS = [
    "id",
    "name",
    "dynasty",
    "museum",
    "material",
    "description",
    "artist",
    "date",
    "object_url",
]

EXPORT_XLSX_COL_WIDTHS: dict[str, float] = {
    "id": 15,
    "name": 40,
    "dynasty": 30,
    "museum": 35,
    "material": 30,
    "description": 60,
    "artist": 30,
    "date": 20,
    "object_url": 50,
}

_XLSX_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF4A3728")
_XLSX_HEADER_FONT = Font(bold=True, color="FFFFFF")
_XLSX_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_XLSX_ROW_WHITE = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
_XLSX_ROW_BEIGE = PatternFill(fill_type="solid", fgColor="FFF5F0E8")

_HTML_TAG_RE = re.compile(r"<[^>]+>", re.IGNORECASE)

NL_SEARCH_PAGE_LIMIT = 10000

MUSEUM_GEO_ENTRIES: list[dict[str, str | float]] = [
    {
        "name": "Metropolitan Museum of Art, New York",
        "lat": 40.7794,
        "lng": -73.9632,
        "country": "United States",
        "city": "New York",
    },
    {
        "name": "Cleveland Museum of Art",
        "lat": 41.5093,
        "lng": -81.6114,
        "country": "United States",
        "city": "Cleveland",
    },
    {
        "name": "Victoria and Albert Museum, London",
        "lat": 51.4966,
        "lng": -0.1722,
        "country": "United Kingdom",
        "city": "London",
    },
    {
        "name": "Art Institute of Chicago",
        "lat": 41.8796,
        "lng": -87.6237,
        "country": "United States",
        "city": "Chicago",
    },
]


class NaturalLanguageQueryBody(BaseModel):
    query: str = Field(..., min_length=1)


app = FastAPI(
    title="Overseas Relic Knowledge Service",
    version="0.1.0",
    description="API for relic metadata: optional Neo4j + sample JSON fallback.",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:5173",
        "http://127.0.0.1:5173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _load_sample_relics() -> list[dict]:
    if not SAMPLE_RELICS_PATH.is_file():
        return []
    with SAMPLE_RELICS_PATH.open(encoding="utf-8") as f:
        return json.load(f)


def _normalize_relic_payload(row: dict) -> dict:
    """Unified API relic shape from JSON (supports legacy ``image``)."""
    r = dict(row)
    legacy_img = r.pop("image", None)
    raw_url = r.get("image_url")
    if isinstance(raw_url, str) and raw_url.strip():
        image_url = raw_url.strip()
    elif isinstance(legacy_img, str) and legacy_img.strip():
        image_url = legacy_img.strip()
    else:
        image_url = ""
    r["image_url"] = image_url
    r["description"] = str(r.get("description") or "").strip()
    r["material"] = str(r.get("material") or "").strip()
    r["dynasty"] = str(r.get("dynasty") or "").strip()
    r["museum"] = str(r.get("museum") or "").strip()
    nm = str(r.get("name") or "").strip()
    r["name"] = nm if nm else "Untitled relic"
    if r.get("id") is not None:
        r["id"] = str(r["id"])
    for key in (
        "artist",
        "date",
        "culture",
        "period",
        "classification",
        "accession_number",
        "dimensions",
        "credit_line",
        "object_url",
        "place",
    ):
        r[key] = str(r.get(key) or "").strip()
    return r


def _filter_sample_relics(
    rows: list[dict],
    dynasty: str | None,
    search: str | None,
    material: str | None = None,
    museum: str | None = None,
) -> list[dict]:
    if dynasty:
        rows = [r for r in rows if (r.get("dynasty") or "") == dynasty]
    if material:
        rows = [r for r in rows if (r.get("material") or "") == material]
    if museum:
        rows = [r for r in rows if (r.get("museum") or "") == museum]
    if search:
        needle = search.lower()
        rows = [
            r
            for r in rows
            if needle in (r.get("name") or "").lower()
            or needle in (r.get("museum") or "").lower()
        ]
    return rows


def _json_filter_facets(
    raw: list[dict],
    d: str | None,
    m: str | None,
    mu: str | None,
    s: str | None,
) -> tuple[list[str], list[str], list[str]]:
    """Distinct dynasty / material / museum for dropdowns given other constraints."""
    dynasty_rows = _filter_sample_relics(raw, None, s, m, mu)
    dynasties = sorted({str(k.get("dynasty")) for k in dynasty_rows if k.get("dynasty")})

    material_rows = _filter_sample_relics(raw, d, s, None, mu)
    materials = sorted({str(k.get("material")) for k in material_rows if k.get("material")})

    museum_rows = _filter_sample_relics(raw, d, s, m, None)
    museums = sorted({str(k.get("museum")) for k in museum_rows if k.get("museum")})

    return dynasties, materials, museums


def _get_sample_relic_by_id(relic_id: str) -> dict | None:
    rid = (relic_id or "").strip()
    if not rid:
        return None
    for row in _load_sample_relics():
        if not isinstance(row, dict):
            continue
        if str(row.get("id") or "").strip() != rid:
            continue
        return _normalize_relic_payload(row)
    return None


def _normalize_list_sort(sort: str | None) -> str:
    s = (sort or "name").strip().lower()
    if s == "period":
        s = "dynasty"
    if s in ("dynasty", "date"):
        return s
    return "name"


def _sort_order_asc(order: str | None) -> bool:
    return (order or "asc").strip().lower() != "desc"


def _sort_sortable_key(row: dict, sort_field: str) -> str:
    if sort_field == "dynasty":
        return (row.get("dynasty") or "").lower()
    if sort_field == "date":
        return (row.get("date") or "").lower()
    return (row.get("name") or "").lower()


def _sort_sample_rows(rows: list[dict], sort_field: str, ascending: bool) -> None:
    reverse = not ascending
    rows.sort(key=lambda r: _sort_sortable_key(r, sort_field), reverse=reverse)


def _stats_from_json() -> dict:
    raw = _load_sample_relics()
    rows = [_normalize_relic_payload(r) for r in raw if isinstance(r, dict)]
    total_relics = len(rows)

    museum_c: Counter[str] = Counter()
    dynasty_c: Counter[str] = Counter()
    material_c: Counter[str] = Counter()

    for r in rows:
        m = str(r.get("museum") or "").strip()
        if m:
            museum_c[m] += 1
        d = str(r.get("dynasty") or "").strip()
        if d:
            dynasty_c[d] += 1
        mat = str(r.get("material") or "").strip()
        if mat:
            material_c[mat] += 1

    by_museum = [{"museum": k, "count": v} for k, v in museum_c.most_common()]
    by_dynasty = [{"dynasty": k, "count": v} for k, v in dynasty_c.most_common(15)]
    by_material = [{"material": k, "count": v} for k, v in material_c.most_common(15)]

    return {
        "total_relics": total_relics,
        "total_museums": len(museum_c),
        "total_dynasties": len(dynasty_c),
        "total_materials": len(material_c),
        "by_museum": by_museum,
        "by_dynasty": by_dynasty,
        "by_material": by_material,
    }


def _related_from_sample(relic_id: str) -> list[dict]:
    """Related relics from JSON: same dynasty OR museum; rank mutual match higher."""
    rid = (relic_id or "").strip()
    if not rid:
        return []
    anchor = _get_sample_relic_by_id(rid)
    if anchor is None:
        return []

    dynasty = anchor.get("dynasty") or ""
    museum = anchor.get("museum") or ""

    ranked: list[tuple[int, dict]] = []
    for raw in _load_sample_relics():
        if not isinstance(raw, dict):
            continue
        oid = str(raw.get("id") or "").strip()
        if oid == rid:
            continue
        rd = raw.get("dynasty") or ""
        rm = raw.get("museum") or ""
        match_d = bool(dynasty and rd == dynasty)
        match_m = bool(museum and rm == museum)
        if not match_d and not match_m:
            continue
        score = 2 if match_d and match_m else 1
        ranked.append((score, _normalize_relic_payload(raw)))

    ranked.sort(key=lambda x: (-x[0], (x[1].get("name") or "").lower()))
    return [row for _, row in ranked[:5]]


def _leading_four_digit_year(date_val: str) -> int | None:
    """Align with Neo4j ``toInteger(left(r.date, 4))``: leading four digits if numeric."""
    s = (date_val or "").strip()
    if len(s) < 4:
        return None
    prefix = s[:4]
    if prefix.isdigit():
        return int(prefix)
    return None


def _matches_advanced_search_row(
    row: dict,
    name: str | None,
    museum: str | None,
    dynasty: str | None,
    material: str | None,
    date_from: int | None,
    date_to: int | None,
    artist: str | None,
    classification: str | None,
) -> bool:
    if name and name.lower() not in (row.get("name") or "").lower():
        return False
    if museum and museum.lower() not in (row.get("museum") or "").lower():
        return False
    if dynasty and (row.get("dynasty") or "") != dynasty:
        return False
    if material and material.lower() not in (row.get("material") or "").lower():
        return False
    if artist and artist.lower() not in (row.get("artist") or "").lower():
        return False
    if classification and classification.lower() not in (row.get("classification") or "").lower():
        return False
    if date_from is not None or date_to is not None:
        y = _leading_four_digit_year(str(row.get("date") or ""))
        if y is None:
            return False
        if date_from is not None and y < date_from:
            return False
        if date_to is not None and y > date_to:
            return False
    return True


def _advanced_search_from_json(
    name: str | None,
    museum: str | None,
    dynasty: str | None,
    material: str | None,
    date_from: int | None,
    date_to: int | None,
    artist: str | None,
    classification: str | None,
    sort_field: str,
    ascending: bool,
    page: int,
    limit: int,
) -> dict:
    raw = _load_sample_relics()
    rows_full = [_normalize_relic_payload(r) for r in raw if isinstance(r, dict)]
    rows_full = [
        r
        for r in rows_full
        if _matches_advanced_search_row(
            r,
            name,
            museum,
            dynasty,
            material,
            date_from,
            date_to,
            artist,
            classification,
        )
    ]
    _sort_sample_rows(rows_full, sort_field, ascending)
    total = len(rows_full)
    skip = (page - 1) * limit
    items = rows_full[skip : skip + limit]
    return {"items": items, "total": total, "page": page, "limit": limit}


def _timeline_from_json() -> dict:
    raw = _load_sample_relics()
    rows = [_normalize_relic_payload(r) for r in raw if isinstance(r, dict)]
    buckets: Counter[int] = Counter()
    for r in rows:
        y = _leading_four_digit_year(str(r.get("date") or ""))
        if y is None:
            continue
        c_ord = y // 100 + 1
        buckets[c_ord] += 1
    periods: list[dict[str, str | int]] = []
    for c_ord in sorted(buckets.keys()):
        century_label, range_label = relics_service.century_ord_labels(c_ord)
        periods.append(
            {"century": century_label, "label": range_label, "count": buckets[c_ord]},
        )
    return {"periods": periods}


def _catalog_export_rows_json_fallback(
    dynasty: str | None,
    search: str | None,
    material: str | None,
    museum: str | None,
    sort_key: str,
    ascending: bool,
) -> list[dict]:
    raw = _load_sample_relics()
    rows_full = _filter_sample_relics(raw, dynasty, search, material, museum)
    rows_full = [_normalize_relic_payload(r) for r in rows_full]
    _sort_sample_rows(rows_full, sort_key, ascending)
    return rows_full[:EXPORT_MAX_ROWS]


def _clean_export_field(raw: Any) -> str:
    """Strip HTML-ish markup and normalize punctuation for CSV / Excel / JSON exports."""
    s = "" if raw is None else str(raw)
    s = _HTML_TAG_RE.sub("", s)
    s = html.unescape(s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    for curly, straight in (
        ("\u201c", '"'),
        ("\u201d", '"'),
        ("\u00ab", '"'),
        ("\u00bb", '"'),
        ("\u2018", "'"),
        ("\u2019", "'"),
        ("\u2032", "'"),
    ):
        s = s.replace(curly, straight)
    for dash in ("\u2013", "\u2014", "\u2212", "\u2010", "\u2011"):
        s = s.replace(dash, "-")
    s = re.sub(r"[ \t]+\n", "\n", s)
    return s.strip()


def _export_row_cells(row: dict) -> list[str]:
    return [_clean_export_field(row.get(k)) for k in EXPORT_CSV_COLUMNS]


def _build_export_csv_string(rows: list[dict]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(EXPORT_CSV_COLUMNS)
    for row in rows:
        writer.writerow(_export_row_cells(row))
    return buf.getvalue()


def _build_export_json_bytes(rows: list[dict]) -> bytes:
    slim = [{k: _clean_export_field(row.get(k)) for k in EXPORT_CSV_COLUMNS} for row in rows]
    return json.dumps(slim, ensure_ascii=False).encode("utf-8")


def _build_export_xlsx_bytes(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Relics"

    for col_idx, key in enumerate(EXPORT_CSV_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = _XLSX_HEADER_FONT
        cell.fill = _XLSX_HEADER_FILL
        cell.alignment = _XLSX_WRAP_TOP
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = EXPORT_XLSX_COL_WIDTHS[key]

    for r_idx, row in enumerate(rows, start=2):
        row_fill = _XLSX_ROW_WHITE if r_idx % 2 == 0 else _XLSX_ROW_BEIGE
        values = _export_row_cells(row)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.alignment = _XLSX_WRAP_TOP
            cell.fill = row_fill

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.get("/stats")
def stats() -> dict:
    """Collection-wide aggregates: Neo4j when available and non-empty, else sample JSON."""
    neo = relics_service.fetch_stats_from_neo4j()
    if neo is not None:
        return neo
    return _stats_from_json()


@app.get("/relics")
def list_relics(
    dynasty: str | None = Query(None, description="Exact match on Relic.dynasty when set."),
    search: str | None = Query(
        None,
        description="Case-insensitive substring match on name or museum.",
    ),
    material: str | None = Query(None, description="Exact match on Relic.material when set."),
    museum: str | None = Query(None, description="Exact match on Relic.museum when set."),
    sort: str | None = Query(
        None,
        description='Sort field: "name", "dynasty", or "date". Legacy "period" maps to dynasty.',
    ),
    order: str | None = Query("asc", description='Sort direction: "asc" or "desc".'),
    page: int = Query(1, ge=1, description="1-based page index."),
    limit: int = Query(10, ge=1, le=100, description="Page size."),
) -> dict:
    """Paginated relics: Neo4j when available, else JSON sample (filters apply before SKIP/LIMIT)."""
    d = (dynasty or "").strip() or None
    s = (search or "").strip() or None
    m = (material or "").strip() or None
    mu = (museum or "").strip() or None
    sort_key = _normalize_list_sort(sort)
    asc = _sort_order_asc(order)

    neo = relics_service.fetch_relics_page_from_neo4j(
        dynasty=d,
        search=s,
        material=m,
        museum=mu,
        page=page,
        limit=limit,
        sort=sort_key,
        ascending=asc,
    )
    if neo is not None:
        return neo

    raw = _load_sample_relics()
    rows_full = _filter_sample_relics(raw, d, s, m, mu)
    rows_full = [_normalize_relic_payload(r) for r in rows_full]
    _sort_sample_rows(rows_full, sort_key, asc)
    total = len(rows_full)
    skip = (page - 1) * limit
    items = rows_full[skip : skip + limit]

    dynasties, materials, museums = _json_filter_facets(raw, d, m, mu, s)

    return {
        "items": items,
        "total": total,
        "page": page,
        "limit": limit,
        "dynasties": dynasties,
        "materials": materials,
        "museums": museums,
    }


@app.get("/relics/search/advanced")
def advanced_search_relics(
    name: str | None = Query(None, description="Substring match on relic name."),
    museum: str | None = Query(None, description="Substring match on museum."),
    dynasty: str | None = Query(None, description="Exact match on dynasty."),
    material: str | None = Query(None, description="Substring match on material."),
    date_from: int | None = Query(
        None,
        description="Include relics whose date starts with a year >= this value.",
    ),
    date_to: int | None = Query(
        None,
        description="Include relics whose date starts with a year <= this value.",
    ),
    artist: str | None = Query(None, description="Substring match on artist."),
    classification: str | None = Query(None, description="Substring match on classification."),
    sort: str | None = Query(
        None,
        description='Sort field: "name", "dynasty", or "date". Legacy "period" maps to dynasty.',
    ),
    order: str | None = Query("asc", description='Sort direction: "asc" or "desc".'),
    page: int = Query(1, ge=1, description="1-based page index."),
    limit: int = Query(10, ge=1, le=100, description="Page size."),
) -> dict:
    """Filtered relic search: Neo4j when available, else JSON sample."""
    q_name = (name or "").strip() or None
    q_museum = (museum or "").strip() or None
    q_dynasty = (dynasty or "").strip() or None
    q_material = (material or "").strip() or None
    q_artist = (artist or "").strip() or None
    q_classification = (classification or "").strip() or None
    sort_key = _normalize_list_sort(sort)
    asc = _sort_order_asc(order)

    neo = relics_service.fetch_advanced_search_from_neo4j(
        name=q_name,
        museum=q_museum,
        dynasty=q_dynasty,
        material=q_material,
        date_from=date_from,
        date_to=date_to,
        artist=q_artist,
        classification=q_classification,
        page=page,
        limit=limit,
        sort=sort_key,
        ascending=asc,
    )
    if neo is not None:
        return neo

    return _advanced_search_from_json(
        q_name,
        q_museum,
        q_dynasty,
        q_material,
        date_from,
        date_to,
        q_artist,
        q_classification,
        sort_key,
        asc,
        page,
        limit,
    )


@app.post("/relics/query/natural")
def natural_language_relic_search(body: NaturalLanguageQueryBody) -> dict:
    """Parse a natural-language query with GPT-4o-mini, then run advanced search."""
    original = body.query.strip()
    if not os.getenv("OPENAI_API_KEY", "").strip():
        raise HTTPException(
            status_code=503,
            detail="OPENAI_API_KEY not configured",
        )

    try:
        parsed = relics_service.extract_nl_filters_with_openai(original)
    except ValueError as e:
        raise HTTPException(status_code=502, detail=str(e)) from e
    except Exception as exc:
        raise HTTPException(
            status_code=502,
            detail="Language model request failed",
        ) from exc

    neo = relics_service.fetch_advanced_search_from_neo4j(
        name=parsed.get("name"),
        museum=parsed.get("museum"),
        dynasty=parsed.get("dynasty"),
        material=parsed.get("material"),
        date_from=parsed.get("date_from"),
        date_to=parsed.get("date_to"),
        artist=parsed.get("artist"),
        classification=None,
        page=1,
        limit=NL_SEARCH_PAGE_LIMIT,
        sort="name",
        ascending=True,
    )
    if neo is not None:
        return {
            "query": body.query,
            "parsed_filters": parsed,
            "items": neo["items"],
            "total": neo["total"],
        }

    fb = _advanced_search_from_json(
        parsed.get("name"),
        parsed.get("museum"),
        parsed.get("dynasty"),
        parsed.get("material"),
        parsed.get("date_from"),
        parsed.get("date_to"),
        parsed.get("artist"),
        None,
        "name",
        True,
        1,
        NL_SEARCH_PAGE_LIMIT,
    )
    return {
        "query": body.query,
        "parsed_filters": parsed,
        "items": fb["items"],
        "total": fb["total"],
    }


@app.get("/relics/export")
async def export_relics_catalog(
    dynasty: str | None = Query(None),
    search: str | None = Query(None),
    material: str | None = Query(None),
    museum: str | None = Query(None),
    sort: str | None = Query(None),
    order: str | None = Query("asc"),
    fmt: Literal["csv", "json", "xlsx"] = Query(..., alias="format"),
) -> StreamingResponse:
    """Export up to 10k relics matching catalog filters as CSV, JSON, or styled XLSX."""
    d = (dynasty or "").strip() or None
    s = (search or "").strip() or None
    m = (material or "").strip() or None
    mu = (museum or "").strip() or None
    sort_key = _normalize_list_sort(sort)
    asc = _sort_order_asc(order)

    neo_rows = relics_service.fetch_relics_export_neo4j(
        dynasty=d,
        search=s,
        material=m,
        museum=mu,
        sort=sort_key,
        ascending=asc,
        max_rows=EXPORT_MAX_ROWS,
    )
    rows = neo_rows if neo_rows is not None else _catalog_export_rows_json_fallback(
        d, s, m, mu, sort_key, asc
    )

    if fmt == "csv":
        csv_body = _build_export_csv_string(rows)

        async def stream_csv():
            yield "\ufeff".encode("utf-8")
            yield csv_body.encode("utf-8")

        return StreamingResponse(
            stream_csv(),
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": 'attachment; filename="relics_export.csv"'},
        )

    if fmt == "json":
        payload = _build_export_json_bytes(rows)

        async def stream_json():
            yield payload

        return StreamingResponse(
            stream_json(),
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="relics_export.json"'},
        )

    payload = _build_export_xlsx_bytes(rows)

    async def stream_xlsx():
        yield payload

    return StreamingResponse(
        stream_xlsx(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="relics_export.xlsx"'},
    )


@app.get("/timeline")
def timeline() -> dict:
    """Relic counts grouped by century from leading date year (Neo4j or JSON)."""
    neo = relics_service.fetch_timeline_from_neo4j()
    if neo is not None:
        return neo
    return _timeline_from_json()


@app.get("/museums/geo")
def museums_geo() -> list[dict]:
    """Known museums with coordinates and collection counts."""
    counts_map = relics_service.get_museum_relic_counts_neo4j()
    if counts_map is None:
        stats = _stats_from_json()
        counts_map = {
            str(row["museum"]): int(row["count"])
            for row in stats.get("by_museum", [])
            if row.get("museum")
        }

    out: list[dict[str, str | float | int]] = []
    for entry in MUSEUM_GEO_ENTRIES:
        name = str(entry["name"])
        out.append(
            {
                "name": name,
                "lat": float(entry["lat"]),
                "lng": float(entry["lng"]),
                "country": str(entry["country"]),
                "city": str(entry["city"]),
                "relic_count": int(counts_map.get(name, 0)),
            },
        )
    return out


@app.get("/relics/{relic_id}")
def get_relic(relic_id: str) -> dict:
    """Single relic (Neo4j when available, otherwise sample JSON)."""
    neo = relics_service.fetch_relic_by_id_from_neo4j(relic_id.strip())
    if neo is not None:
        return neo

    fallback = _get_sample_relic_by_id(relic_id.strip())
    if fallback is None:
        raise HTTPException(status_code=404, detail="Relic not found")
    return fallback


@app.get("/relics/{relic_id}/related")
def related_relics(relic_id: str) -> list[dict]:
    """Related relics from Neo4j when available; JSON sample heuristic when graph is off or yields none."""
    graph_hits = relics_service.fetch_related_relics_from_neo4j(relic_id)
    if graph_hits:
        return graph_hits
    return _related_from_sample(relic_id)
